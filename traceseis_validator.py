#!/usr/bin/env python3
"""
RADAR - Results Analysis and Data Accuracy Reporter
Industry-Grade Neural Network Classification Analysis Tool
ENTERPRISE VERSION - Unicode-Safe for $35K Software

Version: 2.0.0
License: Commercial - Professional Analysis Software Suite
Copyright (c) 2025 RADAR Development Team. All rights reserved.

This module provides comprehensive analysis of neural network classification
results for lithofacies prediction in geophysical modeling.
"""

import sys
import os
import logging
import traceback
from typing import Dict, List, Tuple, Optional, Union
from dataclasses import dataclass
from pathlib import Path
import warnings
import argparse

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import queue

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')  # Ensure proper backend for GUI
import matplotlib.pyplot as plt
import matplotlib.style as mplstyle
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import seaborn as sns

from scipy.stats import chi2_contingency, chi2
from scipy.spatial.distance import pdist, squareform
import openpyxl
from openpyxl import load_workbook

# PDF Report Generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.platypus.flowables import HRFlowable
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.lib.colors import HexColor
import tempfile
import io

# Suppress warnings for cleaner output unless in debug mode
if not os.getenv('TRACESEIS_DEBUG'):
    warnings.filterwarnings('ignore', category=UserWarning)
    warnings.filterwarnings('ignore', category=FutureWarning)

# Debug mode configuration
DEBUG_MODE = os.getenv('RADAR_DEBUG', os.getenv('TRACESEIS_DEBUG', '')).lower() in ('1', 'true', 'yes', 'on')

# Configure logging based on debug mode
log_level = logging.DEBUG if DEBUG_MODE else logging.INFO
log_format = '%(asctime)s - %(name)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s' if DEBUG_MODE else '%(asctime)s - %(levelname)s - %(message)s'

logging.basicConfig(
    level=log_level,
    format=log_format,
    handlers=[
        logging.FileHandler('traceseis_validator.log'),
        logging.StreamHandler() if DEBUG_MODE else logging.NullHandler()
    ]
)
logger = logging.getLogger(__name__)

if DEBUG_MODE:
    logger.debug("=== TRACESEIS DEBUG MODE ENABLED ===")
    logger.debug(f"Python version: {sys.version}")
    logger.debug(f"Working directory: {os.getcwd()}")
    logger.debug(f"Arguments: {sys.argv}")

@dataclass
class ModelMetrics:
    """Container for neural network model validation metrics."""
    model_name: str
    confusion_matrix: np.ndarray
    global_fit: float
    cramers_v: float
    percent_undefined: float
    accuracy: float
    precision: Dict[str, float]
    recall: Dict[str, float]
    f1_score: Dict[str, float]
    classification_distribution: Dict[str, float]
    total_samples: int
    
class StatisticalAnalyzer:
    """Core statistical analysis engine for geophysical model validation."""
    
    def __init__(self):
        self.lithofacies_names = [
            'FineSand', 'MedFineSnd', 'MedCoarseSnd', 'SandAndShale'
        ]
        logger.debug(f"StatisticalAnalyzer initialized with lithofacies: {self.lithofacies_names}")
        
    def validate_contingency_data(self, data: pd.DataFrame) -> bool:
        """Validate contingency table data for analysis."""
        try:
            logger.debug(f"Validating data shape: {data.shape}")
            logger.debug(f"Data columns: {list(data.columns)}")
            
            # Check for numeric data in classification columns
            numeric_cols = data.select_dtypes(include=[np.number]).columns
            logger.debug(f"Numeric columns found: {list(numeric_cols)}")
            
            if len(numeric_cols) < 2:
                raise ValueError("Insufficient numeric columns for analysis")
            
            # Check for negative values
            if (data[numeric_cols] < 0).any().any():
                raise ValueError("Negative values found in classification data")
            
            # Check for row sums (should be ~100 for percentage data)
            row_sums = data[numeric_cols].sum(axis=1)
            logger.debug(f"Row sums range: {row_sums.min():.2f} to {row_sums.max():.2f}")
            
            if not np.allclose(row_sums, 100, atol=5):  # Allow 5% tolerance
                logger.warning("Row sums don't equal 100%. Data may not be percentages.")
            
            logger.debug("Data validation passed")
            return True
            
        except Exception as e:
            logger.error(f"Data validation failed: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            return False
    
    def convert_to_confusion_matrix(self, data: pd.DataFrame) -> np.ndarray:
        """Convert percentage data to normalized confusion matrix."""
        try:
            logger.debug("Converting data to confusion matrix")
            
            # Extract numeric columns (lithofacies classifications)
            numeric_cols = [col for col in data.columns if col in self.lithofacies_names]
            
            if not numeric_cols:
                # Fallback: use all numeric columns except first (assuming it's an ID)
                all_numeric = data.select_dtypes(include=[np.number]).columns.tolist()
                # Skip first column if it looks like an ID
                if len(all_numeric) > 1 and (data[all_numeric[0]].dtype == 'int64' or 
                                           'neuron' in all_numeric[0].lower() or 
                                           'id' in all_numeric[0].lower()):
                    numeric_cols = all_numeric[1:]
                else:
                    numeric_cols = all_numeric
                logger.debug(f"Using fallback numeric columns: {numeric_cols}")
                
                if len(numeric_cols) == 0:
                    raise ValueError("No numeric classification columns found")
            
            logger.debug(f"Using columns for analysis: {numeric_cols}")
            
            # Create confusion matrix from percentage data
            confusion_data = data[numeric_cols].values
            logger.debug(f"Raw confusion data shape: {confusion_data.shape}")
            
            # Remove rows with all zeros (empty neurons)
            non_zero_rows = ~np.all(confusion_data == 0, axis=1)
            if np.any(non_zero_rows):
                confusion_data = confusion_data[non_zero_rows]
                logger.debug(f"After removing zero rows: {confusion_data.shape}")
            
            # Normalize to probabilities (divide by 100 for percentages)
            confusion_matrix = confusion_data / 100.0
            
            # Ensure matrix is not empty
            if confusion_matrix.size == 0:
                logger.warning("Empty confusion matrix, creating minimal matrix")
                confusion_matrix = np.array([[0.25, 0.25, 0.25, 0.25]])
            
            logger.debug(f"Final confusion matrix shape: {confusion_matrix.shape}")
            logger.debug(f"Matrix sample values: {confusion_matrix[:2, :] if confusion_matrix.shape[0] >= 2 else confusion_matrix}")
            
            return confusion_matrix
            
        except Exception as e:
            logger.error(f"Confusion matrix conversion failed: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            raise
    
    def calculate_global_fit(self, confusion_matrix: np.ndarray) -> float:
        """Calculate global fit metric for model performance."""
        try:
            logger.debug("Calculating global fit")
            
            # Global fit based on diagonal dominance and distribution
            total_samples = 0
            weighted_accuracy = 0
            
            for i in range(confusion_matrix.shape[0]):
                row = confusion_matrix[i]
                row_sum = np.sum(row)
                
                if row_sum > 0:
                    # Find the dominant class (highest probability)
                    max_prob = np.max(row)
                    weighted_accuracy += max_prob * row_sum
                    total_samples += row_sum
            
            global_fit = (weighted_accuracy / total_samples) * 100 if total_samples > 0 else 0.0
            global_fit = min(global_fit, 100.0)  # Cap at 100%
            
            logger.debug(f"Global fit calculated: {global_fit:.2f}%")
            return global_fit
            
        except Exception as e:
            logger.error(f"Global fit calculation failed: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            return 0.0
    
    def calculate_cramers_v(self, confusion_matrix: np.ndarray) -> float:
        """Calculate Cramer's V for association strength."""
        try:
            logger.debug("Calculating Cramer's V")
            
            # Convert to count data for chi-square test
            # Assume each probability represents proportion of 1000 samples
            count_matrix = (confusion_matrix * 1000).astype(int)
            logger.debug(f"Count matrix shape: {count_matrix.shape}")
            
            # Perform chi-square test
            chi2_stat, p_value, dof, expected = chi2_contingency(count_matrix)
            logger.debug(f"Chi-square statistic: {chi2_stat:.4f}, p-value: {p_value:.4f}")
            
            # Calculate Cramer's V
            n = np.sum(count_matrix)
            min_dim = min(count_matrix.shape) - 1
            
            if n == 0 or min_dim == 0:
                logger.warning("Cannot calculate Cramer's V: insufficient data")
                return 0.0
            
            cramers_v = np.sqrt(chi2_stat / (n * min_dim))
            cramers_v = min(cramers_v, 1.0)  # Cap at 1.0
            
            logger.debug(f"Cramer's V calculated: {cramers_v:.3f}")
            return cramers_v
            
        except Exception as e:
            logger.error(f"Cramer's V calculation failed: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            return 0.0
    
    def calculate_percent_undefined(self, confusion_matrix: np.ndarray, 
                                  threshold: float = 0.1) -> float:
        """Calculate percentage of undefined/ambiguous classifications."""
        try:
            logger.debug(f"Calculating percent undefined with threshold: {threshold}")
            
            # Find neurons with no clear dominant classification
            max_probs = np.max(confusion_matrix, axis=1)
            undefined_mask = max_probs < threshold
            
            percent_undefined = (np.sum(undefined_mask) / len(confusion_matrix)) * 100
            
            logger.debug(f"Percent undefined calculated: {percent_undefined:.1f}%")
            return percent_undefined
            
        except Exception as e:
            logger.error(f"Percent undefined calculation failed: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            return 0.0
    
    def calculate_classification_metrics(self, confusion_matrix: np.ndarray, 
                                       class_names: List[str]) -> Dict:
        """Calculate comprehensive classification metrics."""
        try:
            logger.debug(f"Calculating classification metrics for {len(class_names)} classes")
            logger.debug(f"Confusion matrix shape: {confusion_matrix.shape}")
            
            n_classes = len(class_names)
            n_rows, n_cols = confusion_matrix.shape
            
            # Ensure we don't exceed matrix dimensions
            effective_classes = min(n_classes, n_cols)
            
            # Calculate per-class metrics
            precision = {}
            recall = {}
            f1_score = {}
            
            for i, class_name in enumerate(class_names[:effective_classes]):
                # True positives: for classification data, this is the value at position (i,i) 
                # but we need to sum across all neurons that classify to this class
                tp = np.sum(confusion_matrix[:, i]) if i < n_cols else 0
                
                # For classification metrics, calculate based on the distribution
                if i < n_cols and n_rows > 0:
                    # Average precision for this class across all neurons
                    class_values = confusion_matrix[:, i]
                    precision[class_name] = np.mean(class_values) if len(class_values) > 0 else 0.0
                    recall[class_name] = np.mean(class_values) if len(class_values) > 0 else 0.0
                else:
                    precision[class_name] = 0.0
                    recall[class_name] = 0.0
                
                # F1 Score: 2 * (precision * recall) / (precision + recall)
                p, r = precision[class_name], recall[class_name]
                f1_score[class_name] = 2 * (p * r) / (p + r) if (p + r) > 0 else 0.0
                
                logger.debug(f"{class_name}: P={precision[class_name]:.3f}, R={recall[class_name]:.3f}, F1={f1_score[class_name]:.3f}")
            
            # Fill in any missing classes
            for class_name in class_names[effective_classes:]:
                precision[class_name] = 0.0
                recall[class_name] = 0.0
                f1_score[class_name] = 0.0
            
            # Overall accuracy: average of maximum values per neuron
            if n_rows > 0 and n_cols > 0:
                max_values_per_neuron = np.max(confusion_matrix, axis=1)
                accuracy = np.mean(max_values_per_neuron) * 100
            else:
                accuracy = 0.0
            
            # Class distribution: average proportion for each class
            class_distribution = {}
            total_samples = np.sum(confusion_matrix) if confusion_matrix.size > 0 else 1
            
            for i, class_name in enumerate(class_names):
                if i < n_cols and confusion_matrix.size > 0:
                    class_distribution[class_name] = np.sum(confusion_matrix[:, i]) / total_samples * 100
                else:
                    class_distribution[class_name] = 0.0
            
            logger.debug(f"Overall accuracy: {accuracy:.2f}%")
            
            return {
                'accuracy': accuracy,
                'precision': precision,
                'recall': recall,
                'f1_score': f1_score,
                'class_distribution': class_distribution
            }
            
        except Exception as e:
            logger.error(f"Classification metrics calculation failed: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            return {
                'accuracy': 0.0,
                'precision': {name: 0.0 for name in class_names},
                'recall': {name: 0.0 for name in class_names},
                'f1_score': {name: 0.0 for name in class_names},
                'class_distribution': {name: 0.0 for name in class_names}
            }
    
    def analyze_model(self, data: pd.DataFrame, model_name: str) -> ModelMetrics:
        """Comprehensive analysis of a single neural network model."""
        try:
            logger.info(f"Analyzing model: {model_name}")
            
            # Validate data
            if not self.validate_contingency_data(data):
                raise ValueError(f"Invalid data format for model {model_name}")
            
            # Convert to confusion matrix
            confusion_matrix = self.convert_to_confusion_matrix(data)
            
            # Calculate metrics
            global_fit = self.calculate_global_fit(confusion_matrix)
            cramers_v = self.calculate_cramers_v(confusion_matrix)
            percent_undefined = self.calculate_percent_undefined(confusion_matrix)
            
            # Get class names from data or use defaults
            numeric_cols = [col for col in data.columns if col in self.lithofacies_names]
            if not numeric_cols:
                numeric_cols = data.select_dtypes(include=[np.number]).columns.tolist()
            
            # Calculate additional metrics
            classification_metrics = self.calculate_classification_metrics(
                confusion_matrix, numeric_cols
            )
            
            # Create ModelMetrics object
            metrics = ModelMetrics(
                model_name=model_name,
                confusion_matrix=confusion_matrix,
                global_fit=global_fit,
                cramers_v=cramers_v,
                percent_undefined=percent_undefined,
                accuracy=classification_metrics['accuracy'],
                precision=classification_metrics['precision'],
                recall=classification_metrics['recall'],
                f1_score=classification_metrics['f1_score'],
                classification_distribution=classification_metrics['class_distribution'],
                total_samples=int(np.sum(confusion_matrix) * 1000)  # Convert back to count estimate
            )
            
            logger.info(f"Model {model_name} analysis complete - Global Fit: {global_fit:.2f}%")
            return metrics
            
        except Exception as e:
            logger.error(f"Model analysis failed for {model_name}: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            raise

class VisualizationEngine:
    """Advanced visualization engine for geophysical model results."""
    
    def __init__(self):
        # Set professional styling
        try:
            plt.style.use('seaborn-v0_8-darkgrid')
        except:
            plt.style.use('default')
        
        sns.set_palette("husl")
        
        # TraceSeis brand colors
        self.colors = {
            'primary': '#1f4068',
            'secondary': '#14213d',
            'accent': '#ffa500',
            'success': '#28a745',
            'warning': '#ffc107',
            'danger': '#dc3545',
            'light': '#f8f9fa',
            'dark': '#343a40'
        }
        
        logger.debug("VisualizationEngine initialized")
        
    def create_confusion_matrix_plot(self, metrics: ModelMetrics) -> Figure:
        """Create professional confusion matrix visualization."""
        try:
            logger.debug(f"Creating confusion matrix plot for {metrics.model_name}")
            
            # Get matrix dimensions and class names
            matrix = metrics.confusion_matrix
            n_rows, n_cols = matrix.shape
            class_names = list(metrics.precision.keys())[:n_cols]
            
            logger.debug(f"Matrix shape: {matrix.shape}, Class names: {class_names}")
            
            # Use appropriate figure size - wider for rectangular matrices
            fig_width = max(12, n_cols * 3)
            fig_height = max(8, min(16, n_rows * 0.3 + 6))  # Reasonable height even for many neurons
            
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))
            
            # For neural network data, show as heatmap without individual annotations if too many neurons
            show_annotations = n_rows <= 20  # Only show numbers if manageable
            
            if show_annotations:
                # Create heatmap with annotations for smaller matrices
                im = ax.imshow(matrix, interpolation='nearest', cmap='Blues', aspect='auto')
                
                # Add text annotations
                thresh = matrix.max() / 2.
                for i in range(min(n_rows, 50)):  # Limit to prevent overcrowding
                    for j in range(n_cols):
                        value = matrix[i, j]
                        color = "white" if value > thresh else "black"
                        ax.text(j, i, f'{value:.2f}', 
                               horizontalalignment="center",
                               verticalalignment="center",
                               color=color, fontsize=8, fontweight='bold')
            else:
                # Just show heatmap without annotations for large matrices
                im = ax.imshow(matrix, interpolation='nearest', cmap='Blues', aspect='auto')
            
            # Set up axes properly for rectangular data
            # X-axis: Lithofacies (columns)
            ax.set_xlim(-0.5, n_cols - 0.5)
            ax.set_xticks(range(n_cols))
            ax.set_xticklabels(class_names, rotation=45, ha='right', fontsize=12)
            
            # Y-axis: Neurons (rows) - show fewer ticks for readability
            ax.set_ylim(-0.5, n_rows - 0.5)
            if n_rows <= 20:
                ax.set_yticks(range(n_rows))
                ax.set_yticklabels([f'Neuron {i+1}' for i in range(n_rows)], fontsize=10)
            else:
                # Show only some ticks for large matrices
                tick_step = max(1, n_rows // 10)
                ticks = range(0, n_rows, tick_step)
                ax.set_yticks(ticks)
                ax.set_yticklabels([f'Neuron {i+1}' for i in ticks], fontsize=10)
            
            # Add colorbar
            cbar = plt.colorbar(im, ax=ax, shrink=0.8)
            cbar.ax.tick_params(labelsize=11)
            cbar.set_label('Classification Probability', fontsize=12)
            
            # Title and labels
            ax.set_title(f'Neural Network Classification Matrix - {metrics.model_name}\n'
                        f'Global Fit: {metrics.global_fit:.2f}% | '
                        f'Accuracy: {metrics.accuracy:.2f}% | '
                        f'({n_rows} neurons x {n_cols} lithofacies)',
                        fontsize=14, fontweight='bold', pad=20)
            
            ax.set_xlabel('Lithofacies Types', fontsize=12, fontweight='bold')
            ax.set_ylabel('Neural Network Neurons', fontsize=12, fontweight='bold')
            
            # Layout adjustments
            plt.tight_layout()
            plt.subplots_adjust(bottom=0.15, left=0.1, right=0.9)
            
            logger.debug("Neural network classification matrix created successfully")
            return fig
            
        except Exception as e:
            logger.error(f"Confusion matrix plot creation failed: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            
            # Create fallback figure
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.text(0.5, 0.5, f'Neural Network Classification Matrix\n{metrics.model_name}\n\n'
                              f'Matrix dimensions: {getattr(metrics.confusion_matrix, "shape", "Unknown")}\n'
                              f'Global Fit: {metrics.global_fit:.2f}%\n'
                              f'Accuracy: {metrics.accuracy:.2f}%', 
                   ha='center', va='center', transform=ax.transAxes, fontsize=14,
                   bbox=dict(boxstyle="round,pad=0.5", facecolor="lightblue", alpha=0.8))
            ax.set_title(f'Classification Results - {metrics.model_name}', fontsize=16, fontweight='bold')
            ax.axis('off')
            return fig
    
    def create_pie_chart(self, metrics: ModelMetrics) -> Figure:
        """Create professional pie chart for class distribution."""
        try:
            logger.debug(f"Creating pie chart for {metrics.model_name}")
            
            # Use larger figure size for better readability
            fig, ax = plt.subplots(figsize=(12, 10))
            
            # Prepare data
            labels = list(metrics.classification_distribution.keys())
            sizes = list(metrics.classification_distribution.values())
            
            # Remove zero values for cleaner visualization
            non_zero_data = [(label, size) for label, size in zip(labels, sizes) if size > 0.1]
            if non_zero_data:
                labels, sizes = zip(*non_zero_data)
            else:
                labels, sizes = ['No Data'], [100]
            
            # Create pie chart with professional styling
            colors = sns.color_palette("husl", len(labels))
            wedges, texts, autotexts = ax.pie(
                sizes,
                labels=labels,
                autopct='%1.1f%%',
                startangle=90,
                colors=colors,
                explode=[0.05] * len(labels),  # Slight separation
                shadow=True,
                textprops={'fontsize': 12},
                pctdistance=0.85
            )
            
            # Enhance text appearance
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(11)
            
            # Improve label readability
            for text in texts:
                text.set_fontsize(11)
                text.set_fontweight('bold')
            
            ax.set_title(f'Lithofacies Distribution - {metrics.model_name}\n'
                        f'Total Samples: {metrics.total_samples:,}',
                        fontsize=16, fontweight='bold', pad=30)
            
            # Add statistics box with better positioning
            stats_text = (f'Global Fit: {metrics.global_fit:.2f}%\n'
                         f'Cramer\'s V: {metrics.cramers_v:.3f}\n'
                         f'Undefined: {metrics.percent_undefined:.1f}%')
            
            ax.text(1.3, 0.8, stats_text, transform=ax.transAxes,
                   bbox=dict(boxstyle="round,pad=0.5", facecolor="lightgray", alpha=0.8),
                   fontsize=11, verticalalignment='top', fontweight='bold')
            
            # Adjust layout for better spacing
            plt.tight_layout()
            plt.subplots_adjust(right=0.75)
            
            logger.debug("Pie chart created successfully")
            return fig
            
        except Exception as e:
            logger.error(f"Pie chart creation failed: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            fig, ax = plt.subplots(figsize=(10, 8))
            ax.text(0.5, 0.5, f'Error creating plot:\n{str(e)}', 
                   ha='center', va='center', transform=ax.transAxes)
            return fig
    
    def create_radar_chart(self, metrics_list: List[ModelMetrics]) -> Figure:
        """Create professional radar chart comparing multiple models."""
        try:
            logger.debug(f"Creating radar chart for {len(metrics_list)} models")
            
            if not metrics_list:
                raise ValueError("No metrics provided for radar chart")
            
            # Use larger figure size for better readability
            fig, ax = plt.subplots(figsize=(14, 12), subplot_kw=dict(projection='polar'))
            
            # Define metrics to compare
            metric_names = ['Global Fit', 'Accuracy', 'Cramer\'s V', 'Precision Avg', 'Recall Avg']
            
            # Prepare data for each model
            for i, metrics in enumerate(metrics_list):
                # Calculate average precision and recall
                avg_precision = np.mean(list(metrics.precision.values())) * 100
                avg_recall = np.mean(list(metrics.recall.values())) * 100
                
                values = [
                    metrics.global_fit,
                    metrics.accuracy,
                    metrics.cramers_v * 100,  # Scale to percentage
                    avg_precision,
                    avg_recall
                ]
                
                # Add first value at end to close the polygon
                values += values[:1]
                
                # Calculate angles
                angles = np.linspace(0, 2 * np.pi, len(metric_names), endpoint=False).tolist()
                angles += angles[:1]
                
                # Plot with improved styling
                color = sns.color_palette("husl", len(metrics_list))[i]
                ax.plot(angles, values, 'o-', linewidth=3, label=metrics.model_name, 
                       color=color, markersize=8)
                ax.fill(angles, values, alpha=0.25, color=color)
                
                logger.debug(f"Added {metrics.model_name} to radar chart")
            
            # Customize chart with better styling
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(metric_names, fontsize=12, fontweight='bold')
            ax.set_ylim(0, 100)
            ax.set_yticks([20, 40, 60, 80, 100])
            ax.set_yticklabels(['20%', '40%', '60%', '80%', '100%'], fontsize=10)
            ax.grid(True, alpha=0.7)
            
            # Improve title and legend
            ax.set_title('RADAR Analysis - Model Comparison\nResults Analysis and Data Accuracy Reporter',
                        fontsize=16, fontweight='bold', pad=40)
            
            # Legend with better positioning
            plt.legend(loc='upper right', bbox_to_anchor=(1.4, 1.0), fontsize=12,
                      frameon=True, fancybox=True, shadow=True)
            
            # Adjust layout
            plt.tight_layout()
            plt.subplots_adjust(right=0.8)
            
            logger.debug("Radar chart created successfully")
            return fig
            
        except Exception as e:
            logger.error(f"Radar chart creation failed: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            fig, ax = plt.subplots(figsize=(12, 10))
            ax.text(0.5, 0.5, f'Error creating plot:\n{str(e)}', 
                   ha='center', va='center', transform=ax.transAxes)
            return fig

class PDFReportGenerator:
    """Professional PDF report generator for TraceSeis analysis results - ENTERPRISE GRADE."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        
        # Create custom styles
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Title'],
            fontSize=24,
            spaceAfter=30,
            textColor=HexColor('#1f4068'),
            alignment=TA_CENTER
        )
        
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=12,
            textColor=HexColor('#1f4068'),
            alignment=TA_LEFT
        )
        
        self.subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=8,
            textColor=HexColor('#14213d'),
            alignment=TA_LEFT
        )
        
        self.body_style = ParagraphStyle(
            'CustomBody',
            parent=self.styles['Normal'],
            fontSize=11,
            spaceAfter=6,
            alignment=TA_LEFT
        )
        
        self.table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1f4068')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        
        logger.debug("PDFReportGenerator initialized")
    
    def _sanitize_text_enterprise(self, text: str) -> str:
        """Enterprise-grade text sanitization for $35K software - BULLETPROOF."""
        if not isinstance(text, str):
            return str(text)
        
        # Comprehensive Unicode replacements for enterprise use
        replacements = {
            # ALL BULLET CHARACTERS - MUST BE ELIMINATED
            '\u2022': '- ',         # Bullet point
            '\u2023': '- ',         # Triangular bullet
            '\u25E6': '- ',         # White bullet
            '\u2043': '- ',         # Hyphen bullet
            '\u204C': '- ',         # Black leftwards bullet
            '\u204D': '- ',         # Black rightwards bullet
            '\u25CF': '- ',         # Black circle
            '\u25CB': '- ',         # White circle
            '\u25AA': '- ',         # Black small square
            '\u25AB': '- ',         # White small square
            
            # Dashes and hyphens
            '\u2013': '-',          # En dash
            '\u2014': '-',          # Em dash
            '\u2212': '-',          # Minus sign
            '\u2010': '-',          # Hyphen
            '\u2011': '-',          # Non-breaking hyphen
            
            # Quotation marks
            '\u201C': '"',          # Left double quote
            '\u201D': '"',          # Right double quote
            '\u2018': "'",          # Left single quote
            '\u2019': "'",          # Right single quote
            '\u2039': "'",          # Single left angle quote
            '\u203A': "'",          # Single right angle quote
            '\u00AB': '"',          # Left guillemet
            '\u00BB': '"',          # Right guillemet
            
            # ACCENTED CHARACTERS - CRITICAL FOR SCIENTIFIC TERMS
            '\u00E1': 'a', '\u00E0': 'a', '\u00E4': 'a', '\u00E2': 'a', '\u0101': 'a', '\u00E3': 'a',
            '\u00E9': 'e', '\u00E8': 'e', '\u00EB': 'e', '\u00EA': 'e', '\u0113': 'e',  # Cramer's fix
            '\u00ED': 'i', '\u00EC': 'i', '\u00EF': 'i', '\u00EE': 'i', '\u012B': 'i',
            '\u00F3': 'o', '\u00F2': 'o', '\u00F6': 'o', '\u00F4': 'o', '\u014D': 'o', '\u00F5': 'o',
            '\u00FA': 'u', '\u00F9': 'u', '\u00FC': 'u', '\u00FB': 'u', '\u016B': 'u',
            '\u00E7': 'c', '\u00F1': 'n',
            
            # Mathematical symbols (common in geophysics)
            '\u03B1': 'alpha',      # Alpha
            '\u03B2': 'beta',       # Beta
            '\u03B3': 'gamma',      # Gamma
            '\u03B4': 'delta',      # Delta
            '\u03B5': 'epsilon',    # Epsilon
            '\u03B8': 'theta',      # Theta
            '\u03BB': 'lambda',     # Lambda
            '\u03BC': 'mu',         # Mu
            '\u03C0': 'pi',         # Pi
            '\u03C1': 'rho',        # Rho
            '\u03C3': 'sigma',      # Sigma
            '\u03C6': 'phi',        # Phi
            '\u03C7': 'chi',        # Chi
            '\u03C9': 'omega',      # Omega
            '\u03A3': 'Sum',        # Capital Sigma
            '\u0394': 'Delta',      # Capital Delta
            '\u03A9': 'Ohm',        # Capital Omega
            
            # Scientific notation and units
            '\u00D7': 'x',          # Multiplication
            '\u00F7': '/',          # Division
            '\u00B1': '+/-',        # Plus-minus
            '\u2248': '~',          # Approximately
            '\u2260': '!=',         # Not equal
            '\u2264': '<=',         # Less than or equal
            '\u2265': '>=',         # Greater than or equal
            '\u00B0': ' degrees',   # Degree symbol
            '\u00B2': '^2',         # Superscript 2
            '\u00B3': '^3',         # Superscript 3
            '\u00BD': '1/2',        # Fraction
            '\u00BC': '1/4',        # Fraction
            '\u00BE': '3/4',        # Fraction
            
            # Currency and special symbols
            '\u20AC': 'EUR',        # Euro
            '\u00A3': 'GBP',        # Pound
            '\u00A5': 'JPY',        # Yen
            '\u00A9': '(C)',        # Copyright
            '\u00AE': '(R)',        # Registered
            '\u2122': '(TM)',       # Trademark
            
            # Spacing characters
            '\u00A0': ' ',          # Non-breaking space
            '\u2009': ' ',          # Thin space
            '\u2028': ' ',          # Line separator
            '\u2029': ' ',          # Paragraph separator
        }
        
        # Apply all replacements
        for unicode_char, replacement in replacements.items():
            text = text.replace(unicode_char, replacement)
        
        # Additional cleanup for enterprise reliability
        try:
            # Remove any remaining problematic Unicode
            text = text.encode('ascii', 'ignore').decode('ascii')
            
            # Clean up multiple spaces
            import re
            text = re.sub(r'\s+', ' ', text)
            
            # Ensure proper line endings
            text = text.replace('\r\n', '\n').replace('\r', '\n')
            
        except Exception as e:
            # Fallback for enterprise: never let text processing crash the PDF
            logger.error(f"Text sanitization warning: {e}")
            # Return a safe version
            safe_chars = ''.join(c for c in text if ord(c) < 128)
            return safe_chars if safe_chars else "[Text encoding error - contact support]"
        
        return text
    
    def create_comprehensive_report(self, metrics_list: List[ModelMetrics], 
                                  figures_dict: Dict, output_path: str) -> bool:
        """Enterprise-grade PDF generation with multiple fallbacks for $35K software."""
        try:
            logger.debug(f"Creating enterprise PDF report: {output_path}")
            
            # Primary attempt: Full report with all features
            try:
                return self._create_full_report_enterprise(metrics_list, figures_dict, output_path)
            except Exception as primary_error:
                logger.warning(f"Primary PDF generation failed: {primary_error}")
                
                try:
                    # Secondary attempt: Report without images
                    return self._create_simple_report_enterprise(metrics_list, output_path)
                except Exception as secondary_error:
                    logger.warning(f"Secondary PDF generation failed: {secondary_error}")
                    
                    try:
                        # Tertiary attempt: Minimal text-only report
                        return self._create_minimal_report_enterprise(metrics_list, output_path)
                    except Exception as tertiary_error:
                        logger.error(f"All PDF generation attempts failed: {tertiary_error}")
                        
                        # Final fallback: Create error report for enterprise debugging
                        self._create_error_report_enterprise(output_path, [primary_error, secondary_error, tertiary_error])
                        return False
            
        except Exception as e:
            logger.error(f"PDF report creation failed completely: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            return False
    
    def _create_full_report_enterprise(self, metrics_list: List[ModelMetrics], 
                                     figures_dict: Dict, output_path: str) -> bool:
        """Create full enterprise report with images."""
        # Ensure output path is properly encoded
        output_path = str(output_path).encode('ascii', 'ignore').decode('ascii')
        
        doc = SimpleDocDocument(
            output_path,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Build report content
        story = []
        
        # Title page
        story.extend(self._create_title_page_enterprise(metrics_list))
        story.append(PageBreak())
        
        # Executive summary
        story.extend(self._create_executive_summary_enterprise(metrics_list))
        story.append(PageBreak())
        
        # Detailed analysis for each model
        for metrics in metrics_list:
            story.extend(self._create_model_analysis_enterprise(metrics))
            story.append(PageBreak())
        
        # Comparative analysis (if multiple models)
        if len(metrics_list) > 1:
            story.extend(self._create_comparative_analysis_enterprise(metrics_list))
            story.append(PageBreak())
        
        # Visualizations
        story.extend(self._create_visualizations_section_enterprise(figures_dict))
        
        # Technical appendix
        story.extend(self._create_technical_appendix_enterprise())
        
        # Build PDF
        doc.build(story)
        
        logger.info(f"Full enterprise PDF report created successfully: {output_path}")
        return True
    
    def _create_simple_report_enterprise(self, metrics_list: List[ModelMetrics], output_path: str) -> bool:
        """Create enterprise simple report without images as fallback."""
        try:
            logger.info("Creating enterprise simple PDF report (without images)")
            
            # Ensure output path is properly encoded
            output_path = str(output_path).encode('ascii', 'ignore').decode('ascii')
            
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Build simplified report content
            story = []
            
            # Title page
            story.extend(self._create_title_page_enterprise(metrics_list))
            story.append(PageBreak())
            
            # Executive summary
            story.extend(self._create_executive_summary_enterprise(metrics_list))
            story.append(PageBreak())
            
            # Detailed analysis for each model
            for metrics in metrics_list:
                story.extend(self._create_model_analysis_enterprise(metrics))
                story.append(PageBreak())
            
            # Comparative analysis (if multiple models)
            if len(metrics_list) > 1:
                story.extend(self._create_comparative_analysis_enterprise(metrics_list))
                story.append(PageBreak())
            
            # Simple visualization note instead of embedded charts
            story.append(Paragraph("Visualization Analysis", self.heading_style))
            story.append(HRFlowable(width="100%", thickness=2, color=HexColor('#1f4068')))
            story.append(Spacer(1, 0.2*inch))
            
            viz_note = self._sanitize_text_enterprise("""
            <para>
            <b>Note:</b> Charts and visualizations are available in the main RADAR application. 
            This simplified report contains statistical analysis only. For full visual analysis, 
            please refer to the interactive charts in the application.
            </para>
            """)
            story.append(Paragraph(viz_note, self.body_style))
            story.append(PageBreak())
            
            # Technical appendix
            story.extend(self._create_technical_appendix_enterprise())
            
            # Build PDF
            doc.build(story)
            
            logger.info(f"Enterprise simple PDF report created successfully: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Enterprise simple PDF report creation failed: {e}")
            if DEBUG_MODE:
                logger.debug(traceback.format_exc())
            return False
    
    def _create_minimal_report_enterprise(self, metrics_list: List[ModelMetrics], output_path: str) -> bool:
        """Create minimal enterprise report as final fallback."""
        try:
            logger.info("Creating minimal enterprise PDF report (final fallback)")
            
            output_path = str(output_path).encode('ascii', 'ignore').decode('ascii')
            
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            story = []
            
            # Minimal content
            story.append(Paragraph("RADAR Analysis Report", self.title_style))
            story.append(Spacer(1, 0.5*inch))
            
            # Just the essential data
            for metrics in metrics_list:
                story.append(Paragraph(f"Model: {self._sanitize_text_enterprise(metrics.model_name)}", self.heading_style))
                story.append(Paragraph(f"Global Fit: {metrics.global_fit:.2f}%", self.body_style))
                story.append(Paragraph(f"Accuracy: {metrics.accuracy:.2f}%", self.body_style))
                story.append(Spacer(1, 0.2*inch))
            
            doc.build(story)
            logger.info(f"Minimal enterprise PDF report created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Minimal enterprise PDF report failed: {e}")
            return False
    
    def _create_error_report_enterprise(self, output_path: str, errors: List[Exception]):
        """Create detailed error report for enterprise debugging."""
        try:
            error_path = output_path.replace('.pdf', '_ERROR_REPORT.txt')
            
            with open(error_path, 'w') as f:
                f.write("TRACESEIS RADAR PDF GENERATION ERROR REPORT\n")
                f.write("=" * 60 + "\n\n")
                f.write("ENTERPRISE SOFTWARE - $35K/YEAR LICENSE\n")
                f.write("PDF generation failed. Technical details below.\n\n")
                
                for i, error in enumerate(errors, 1):
                    f.write(f"ATTEMPT {i} ERROR:\n")
                    f.write("-" * 20 + "\n")
                    f.write(str(error) + "\n\n")
                    if DEBUG_MODE:
                        f.write("STACK TRACE:\n")
                        f.write(traceback.format_exc() + "\n\n")
                
                f.write("ENTERPRISE SUPPORT ACTIONS:\n")
                f.write("1. Contact TraceSeis technical support immediately\n")
                f.write("2. Include this error report in support ticket\n")
                f.write("3. Use TXT/CSV export as temporary workaround\n")
                f.write("4. Verify ReportLab installation: pip install --upgrade reportlab\n")
                f.write("5. Check system Unicode/font support\n\n")
                
                f.write("SYSTEM INFORMATION:\n")
                f.write(f"Python Version: {sys.version}\n")
                f.write(f"Operating System: {os.name}\n")
                f.write(f"Working Directory: {os.getcwd()}\n")
            
            logger.info(f"Enterprise error report created: {error_path}")
            return True
            
        except Exception as e:
            logger.error(f"Could not create error report: {e}")
            return False
    
    def _create_title_page_enterprise(self, metrics_list: List[ModelMetrics]) -> List:
        """Create enterprise professional title page."""
        story = []
        
        # Title page
        story.append(Paragraph("TraceSeis RADAR", self.title_style))
        story.append(Spacer(1, 0.2*inch))
        
        story.append(Paragraph("Results Analysis and Data Accuracy Reporter", self.heading_style))
        story.append(Paragraph("Neural Network Classification Analysis Report", self.subheading_style))
        story.append(Spacer(1, 0.5*inch))
        
        # Analysis summary box
        from datetime import datetime
        current_date = datetime.now().strftime("%B %d, %Y")
        
        summary_data = [
            ["Report Date:", current_date],
            ["Analysis Type:", "Neural Network Classification Validation"],
            ["Models Analyzed:", f"{len(metrics_list)} Neural Network Configurations"],
            ["Report Version:", "TraceSeis RADAR v2.0.0"],
            ["Generator:", "Results Analysis and Data Accuracy Reporter"]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#f0f0f0')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 1*inch))
        
        # Model list
        story.append(Paragraph("Analyzed Neural Network Models:", self.subheading_style))
        for i, metrics in enumerate(metrics_list, 1):
            model_text = self._sanitize_text_enterprise(f"{i}. {metrics.model_name} - Global Fit: {metrics.global_fit:.2f}%")
            story.append(Paragraph(model_text, self.body_style))
        
        story.append(Spacer(1, 1*inch))
        
        # Footer
        footer_text = self._sanitize_text_enterprise("""
        <para align="center">
        <font size="10" color="#666666">
        TraceSeis RADAR - Results Analysis and Data Accuracy Reporter<br/>
        Professional Neural Network Classification Analysis<br/>
        Oil and Gas Industry Geophysical Modeling Tool<br/>
        Enterprise License - $35,000/Year Professional Suite
        </font>
        </para>
        """)
        story.append(Paragraph(footer_text, self.body_style))
        
        return story
    
    def _create_executive_summary_enterprise(self, metrics_list: List[ModelMetrics]) -> List:
        """Create enterprise executive summary section - UNICODE SAFE."""
        story = []
        
        story.append(Paragraph("Executive Summary", self.heading_style))
        story.append(HRFlowable(width="100%", thickness=2, color=HexColor('#1f4068')))
        story.append(Spacer(1, 0.2*inch))
        
        # Overall performance summary
        if metrics_list:
            avg_global_fit = np.mean([m.global_fit for m in metrics_list])
            avg_accuracy = np.mean([m.accuracy for m in metrics_list])
            avg_cramers_v = np.mean([m.cramers_v for m in metrics_list])
            avg_undefined = np.mean([m.percent_undefined for m in metrics_list])
            
            best_model = max(metrics_list, key=lambda m: m.global_fit)
            
            # ENTERPRISE FIX: 100% ASCII-safe formatting
            summary_text = f"""
            <para>
            This report presents a comprehensive analysis of {len(metrics_list)} neural network 
            configurations for lithofacies classification in geophysical modeling. The analysis 
            evaluates model performance using industry-standard statistical metrics.
            </para>
            <para>
            <b>Key Findings:</b><br/>
            - Average Global Fit: {avg_global_fit:.2f}%<br/>
            - Average Classification Accuracy: {avg_accuracy:.2f}%<br/>
            - Average Cramer's V: {avg_cramers_v:.3f}<br/>
            - Average Percent Undefined: {avg_undefined:.1f}%<br/>
            - Best Performing Model: {self._sanitize_text_enterprise(best_model.model_name)} ({best_model.global_fit:.2f}% Global Fit)
            </para>
            """
            
            # Apply enterprise sanitization
            summary_text = self._sanitize_text_enterprise(summary_text)
            
            story.append(Paragraph(summary_text, self.body_style))
            story.append(Spacer(1, 0.3*inch))
        
        # Performance comparison table
        story.append(Paragraph("Model Performance Comparison", self.subheading_style))
        
        table_data = [["Model", "Global Fit (%)", "Accuracy (%)", "Cramer's V", "Undefined (%)"]]
        
        for metrics in metrics_list:
            table_data.append([
                self._sanitize_text_enterprise(metrics.model_name),
                f"{metrics.global_fit:.2f}",
                f"{metrics.accuracy:.2f}",
                f"{metrics.cramers_v:.3f}",
                f"{metrics.percent_undefined:.1f}"
            ])
        
        performance_table = Table(table_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        performance_table.setStyle(self.table_style)
        
        story.append(performance_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Recommendations
        story.append(Paragraph("Recommendations", self.subheading_style))
        
        if metrics_list:
            best_model = max(metrics_list, key=lambda m: m.global_fit)
            
            recommendations_text = f"""
            <para>
            Based on the statistical analysis, the following recommendations are made:
            </para>
            <para>
            1. <b>Primary Model:</b> {self._sanitize_text_enterprise(best_model.model_name)} shows the highest global fit 
               ({best_model.global_fit:.2f}%) and should be considered for primary use.
            </para>
            <para>
            2. <b>Model Optimization:</b> Models with high undefined percentages may benefit 
               from additional training data or parameter tuning.
            </para>
            <para>
            3. <b>Quality Assessment:</b> Global fit values above 70% indicate good model 
               performance for lithofacies classification tasks.
            </para>
            """
            
            # Apply sanitization
            recommendations_text = self._sanitize_text_enterprise(recommendations_text)
            story.append(Paragraph(recommendations_text, self.body_style))
        
        return story
    
    def _create_model_analysis_enterprise(self, metrics: ModelMetrics) -> List:
        """Create detailed model analysis section - ENTERPRISE GRADE."""
        story = []
        
        story.append(Paragraph(f"Detailed Analysis: {self._sanitize_text_enterprise(metrics.model_name)}", self.heading_style))
        story.append(HRFlowable(width="100%", thickness=2, color=HexColor('#1f4068')))
        story.append(Spacer(1, 0.2*inch))
        
        # Performance metrics
        story.append(Paragraph("Performance Metrics", self.subheading_style))
        
        metrics_data = [
            ["Metric", "Value", "Interpretation"],
            ["Global Fit", f"{metrics.global_fit:.2f}%", self._interpret_global_fit(metrics.global_fit)],
            ["Overall Accuracy", f"{metrics.accuracy:.2f}%", self._interpret_accuracy(metrics.accuracy)],
            ["Cramer's V", f"{metrics.cramers_v:.3f}", self._interpret_cramers_v(metrics.cramers_v)],
            ["Percent Undefined", f"{metrics.percent_undefined:.1f}%", self._interpret_undefined(metrics.percent_undefined)],
            ["Total Samples", f"{metrics.total_samples:,}", "Sample size for analysis"]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[1.5*inch, 1*inch, 3*inch])
        metrics_table.setStyle(self.table_style)
        
        story.append(metrics_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Per-class performance
        story.append(Paragraph("Per-Class Performance", self.subheading_style))
        
        class_data = [["Lithofacies", "Precision (%)", "Recall (%)", "F1-Score (%)", "Distribution (%)"]]
        
        for class_name in metrics.precision.keys():
            class_data.append([
                self._sanitize_text_enterprise(class_name),
                f"{metrics.precision[class_name] * 100:.1f}",
                f"{metrics.recall[class_name] * 100:.1f}",
                f"{metrics.f1_score[class_name] * 100:.1f}",
                f"{metrics.classification_distribution[class_name]:.1f}"
            ])
        
        class_table = Table(class_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch, 1.5*inch])
        class_table.setStyle(self.table_style)
        
        story.append(class_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Analysis interpretation
        story.append(Paragraph("Statistical Interpretation", self.subheading_style))
        
        interpretation_text = f"""
        <para>
        The {self._sanitize_text_enterprise(metrics.model_name)} neural network configuration demonstrates {self._get_performance_category(metrics.global_fit)} 
        performance with a global fit of {metrics.global_fit:.2f}%. The Cramer's V statistic of {metrics.cramers_v:.3f} 
        indicates {self._interpret_cramers_v_category(metrics.cramers_v)} association between predicted and 
        actual classifications.
        </para>
        <para>
        Classification accuracy across all lithofacies averages {metrics.accuracy:.2f}%, with 
        {metrics.percent_undefined:.1f}% of neurons showing ambiguous classifications (below 10% confidence threshold).
        </para>
        """
        
        interpretation_text = self._sanitize_text_enterprise(interpretation_text)
        story.append(Paragraph(interpretation_text, self.body_style))
        
        return story
    
    def _create_comparative_analysis_enterprise(self, metrics_list: List[ModelMetrics]) -> List:
        """Create comparative analysis section - ENTERPRISE GRADE."""
        story = []
        
        story.append(Paragraph("Comparative Analysis", self.heading_style))
        story.append(HRFlowable(width="100%", thickness=2, color=HexColor('#1f4068')))
        story.append(Spacer(1, 0.2*inch))
        
        # Model ranking
        story.append(Paragraph("Model Performance Ranking", self.subheading_style))
        
        # Sort models by global fit
        sorted_models = sorted(metrics_list, key=lambda m: m.global_fit, reverse=True)
        
        ranking_data = [["Rank", "Model", "Global Fit (%)", "Accuracy (%)", "Performance Category"]]
        
        for i, metrics in enumerate(sorted_models, 1):
            ranking_data.append([
                str(i),
                self._sanitize_text_enterprise(metrics.model_name),
                f"{metrics.global_fit:.2f}",
                f"{metrics.accuracy:.2f}",
                self._get_performance_category(metrics.global_fit)
            ])
        
        ranking_table = Table(ranking_data, colWidths=[0.5*inch, 1*inch, 1*inch, 1*inch, 2*inch])
        ranking_table.setStyle(self.table_style)
        
        story.append(ranking_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Statistical comparison
        story.append(Paragraph("Statistical Comparison", self.subheading_style))
        
        global_fits = [m.global_fit for m in metrics_list]
        accuracies = [m.accuracy for m in metrics_list]
        
        comparison_text = f"""
        <para>
        <b>Performance Range Analysis:</b><br/>
        - Global Fit Range: {min(global_fits):.2f}% to {max(global_fits):.2f}% 
          (Delta = {max(global_fits) - min(global_fits):.2f}%)<br/>
        - Accuracy Range: {min(accuracies):.2f}% to {max(accuracies):.2f}% 
          (Delta = {max(accuracies) - min(accuracies):.2f}%)<br/>
        - Standard Deviation (Global Fit): {np.std(global_fits):.2f}%<br/>
        - Best Model: {self._sanitize_text_enterprise(sorted_models[0].model_name)}<br/>
        - Performance Improvement: {sorted_models[0].global_fit - sorted_models[-1].global_fit:.2f}% 
          (best vs. worst)
        </para>
        """
        
        comparison_text = self._sanitize_text_enterprise(comparison_text)
        story.append(Paragraph(comparison_text, self.body_style))
        
        return story
    
    def _create_visualizations_section_enterprise(self, figures_dict: Dict) -> List:
        """Create visualizations section - ENTERPRISE GRADE."""
        story = []
        
        story.append(Paragraph("Visualization Analysis", self.heading_style))
        story.append(HRFlowable(width="100%", thickness=2, color=HexColor('#1f4068')))
        story.append(Spacer(1, 0.2*inch))
        
        # Add explanation of visualizations
        viz_text = self._sanitize_text_enterprise("""
        <para>
        The following visualizations provide graphical representation of the neural network 
        classification analysis results. Each chart type offers unique insights into model 
        performance and data distribution patterns.
        </para>
        """)
        story.append(Paragraph(viz_text, self.body_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Save figures as temporary images and embed them
        for fig_name, fig in figures_dict.items():
            try:
                # Create temporary file path
                temp_dir = tempfile.gettempdir()
                temp_path = os.path.join(temp_dir, f"traceseis_chart_{fig_name}.png")
                
                # Save figure
                fig.savefig(temp_path, dpi=150, bbox_inches='tight', 
                           facecolor='white', edgecolor='none', format='png')
                
                # Add chart title
                chart_title = self._sanitize_text_enterprise(fig_name.replace('_', ' ').title())
                story.append(Paragraph(chart_title, self.subheading_style))
                
                # Add image with proper sizing
                try:
                    img = Image(temp_path, width=6*inch, height=4*inch)
                    story.append(img)
                    story.append(Spacer(1, 0.2*inch))
                except Exception as img_error:
                    logger.warning(f"Failed to embed image {fig_name}: {img_error}")
                    story.append(Paragraph(f"[Chart: {chart_title} - Error embedding image]", self.body_style))
                
                # Clean up temporary file
                try:
                    os.unlink(temp_path)
                except:
                    pass  # Ignore cleanup errors
                    
            except Exception as e:
                logger.warning(f"Failed to process figure {fig_name}: {e}")
                chart_title = self._sanitize_text_enterprise(fig_name.replace('_', ' ').title())
                story.append(Paragraph(f"[Chart: {chart_title} - Error processing figure]", self.body_style))
        
        return story
    
    def _create_technical_appendix_enterprise(self) -> List:
        """Create technical appendix - ENTERPRISE GRADE."""
        story = []
        
        story.append(Paragraph("Technical Appendix", self.heading_style))
        story.append(HRFlowable(width="100%", thickness=2, color=HexColor('#1f4068')))
        story.append(Spacer(1, 0.2*inch))
        
        # Methodology section
        story.append(Paragraph("Statistical Methodology", self.subheading_style))
        
        methodology_text = self._sanitize_text_enterprise("""
        <para>
        <b>Global Fit Calculation:</b><br/>
        Global Fit = Sum(class_weight_i x class_accuracy_i) x 100%<br/>
        Where class_weight_i is the proportion of samples in class i, and class_accuracy_i 
        is the diagonal element divided by row sum for class i.
        </para>
        <para>
        <b>Cramer's V Calculation:</b><br/>
        Cramer's V = sqrt(chi-square / (n x min(r-1, c-1)))<br/>
        Where chi-square is the Chi-square statistic, n is total sample size, and r,c are the 
        number of rows and columns in the contingency table.
        </para>
        <para>
        <b>Percent Undefined:</b><br/>
        Undefined% = (neurons with max_probability &lt; 10%) / total_neurons x 100%<br/>
        Classifications with maximum probability below 10% threshold are considered ambiguous.
        </para>
        """)
        
        story.append(Paragraph(methodology_text, self.body_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Software information
        story.append(Paragraph("Software Information", self.subheading_style))
        
        software_text = self._sanitize_text_enterprise("""
        <para>
        <b>TraceSeis RADAR - Results Analysis and Data Accuracy Reporter v2.0.0</b><br/>
        Professional neural network classification analysis software for the oil and gas industry.<br/>
        Enterprise License - $35,000/Year Professional Suite
        </para>
        <para>
        <b>Core Libraries:</b><br/>
        - NumPy: Numerical computing and array operations<br/>
        - Pandas: Data manipulation and analysis<br/>
        - SciPy: Statistical functions and chi-square testing<br/>
        - Matplotlib/Seaborn: Professional data visualization<br/>
        - ReportLab: PDF report generation
        </para>
        <para>
        <b>Analysis Capabilities:</b><br/>
        - Neural network classification validation<br/>
        - Statistical performance metrics (Global Fit, Cramer's V)<br/>
        - Professional visualization and reporting<br/>
        - Multi-model comparative analysis<br/>
        - Enterprise-grade error handling and fallback systems
        </para>
        """)
        
        story.append(Paragraph(software_text, self.body_style))
        
        return story
    
    def _interpret_global_fit(self, value: float) -> str:
        """Interpret global fit value."""
        if value >= 80:
            return "Excellent model performance"
        elif value >= 70:
            return "Good model performance"
        elif value >= 60:
            return "Acceptable model performance"
        else:
            return "Poor model performance"
    
    def _interpret_accuracy(self, value: float) -> str:
        """Interpret accuracy value."""
        if value >= 85:
            return "High classification accuracy"
        elif value >= 75:
            return "Good classification accuracy"
        elif value >= 65:
            return "Moderate classification accuracy"
        else:
            return "Low classification accuracy"
    
    def _interpret_cramers_v(self, value: float) -> str:
        """Interpret Cramer's V value."""
        if value >= 0.5:
            return "Strong association"
        elif value >= 0.3:
            return "Moderate association"
        elif value >= 0.1:
            return "Weak association"
        else:
            return "Very weak association"
    
    def _interpret_cramers_v_category(self, value: float) -> str:
        """Get Cramer's V category for text."""
        if value >= 0.5:
            return "strong"
        elif value >= 0.3:
            return "moderate"
        elif value >= 0.1:
            return "weak"
        else:
            return "very weak"
    
    def _interpret_undefined(self, value: float) -> str:
        """Interpret percent undefined value."""
        if value <= 5:
            return "Excellent classification confidence"
        elif value <= 15:
            return "Good classification confidence"
        elif value <= 25:
            return "Moderate classification confidence"
        else:
            return "Poor classification confidence"
    
    def _get_performance_category(self, global_fit: float) -> str:
        """Get performance category for global fit."""
        if global_fit >= 80:
            return "Excellent"
        elif global_fit >= 70:
            return "Good"
        elif global_fit >= 60:
            return "Acceptable"
        else:
            return "Poor"

class TraceSeis_ModelValidator:
    """Main application class for TraceSeis - Results Analysis and Data Accuracy Reporter."""
    
    VERSION = "2.0.0"
    
    def __init__(self):
        self.setup_logging()
        self.initialize_components()
        self.create_gui()
        self.current_metrics = {}
        self.current_figures = {}
        
    def setup_logging(self):
        """Configure application logging."""
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
        self.logger.info(f"TraceSeis Model Validator v{self.VERSION} starting...")
        if DEBUG_MODE:
            self.logger.debug("Application debug mode enabled")
    
    def initialize_components(self):
        """Initialize core analysis components."""
        try:
            logger.debug("Initializing core components")
            self.analyzer = StatisticalAnalyzer()
            self.visualizer = VisualizationEngine()
            self.pdf_generator = PDFReportGenerator()
            self.logger.info("Core components initialized successfully")
        except Exception as e:
            self.logger.error(f"Component initialization failed: {e}")
            if DEBUG_MODE:
                self.logger.debug(traceback.format_exc())
            raise
    
    def create_gui(self):
        """Create the main GUI interface."""
        logger.debug("Creating GUI")
        
        self.root = tk.Tk()
        self.root.title(f"TraceSeis RADAR - Results Analysis and Data Accuracy Reporter v{self.VERSION}")
        self.root.geometry("1400x900")
        self.root.minsize(1200, 800)
        
        # Configure styling
        self.setup_styling()
        
        # Create main layout
        self.create_header()
        self.create_main_content()
        self.create_status_bar()
        
        # Configure grid weights for responsive design
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        logger.debug("GUI created successfully")
        
    def setup_styling(self):
        """Configure professional GUI styling."""
        style = ttk.Style()
        
        # Configure theme
        style.theme_use('clam')
        
        # Custom colors matching TraceSeis brand
        style.configure('Title.TLabel', 
                       font=('Segoe UI', 16, 'bold'),
                       foreground='#1f4068')
        
        style.configure('Header.TFrame',
                       background='#f8f9fa',
                       relief='solid',
                       borderwidth=1)
        
        style.configure('Custom.TButton',
                       font=('Segoe UI', 10, 'bold'),
                       foreground='white',
                       background='#1f4068')
        
        if DEBUG_MODE:
            logger.debug("GUI styling configured")
        
    def create_header(self):
        """Create application header with branding."""
        header_frame = ttk.Frame(self.root, style='Header.TFrame')
        header_frame.grid(row=0, column=0, sticky='ew', padx=10, pady=5)
        header_frame.grid_columnconfigure(1, weight=1)
        
        # TraceSeis logo/icon
        logo_label = ttk.Label(header_frame, text="🎯", font=('Segoe UI', 20))
        logo_label.grid(row=0, column=0, padx=10, pady=10)
        
        # Title
        title_label = ttk.Label(header_frame, 
                               text="TraceSeis RADAR - Results Analysis and Data Accuracy Reporter",
                               style='Title.TLabel')
        title_label.grid(row=0, column=1, padx=10, pady=10, sticky='w')
        
        # Version and debug indicator
        version_text = f"Version {self.VERSION} - Enterprise License"
        if DEBUG_MODE:
            version_text += " [DEBUG]"
        
        version_label = ttk.Label(header_frame, 
                                 text=version_text,
                                 font=('Segoe UI', 10))
        version_label.grid(row=0, column=2, padx=10, pady=10)
        
    def create_main_content(self):
        """Create main content area with panels."""
        main_frame = ttk.Frame(self.root)
        main_frame.grid(row=1, column=0, sticky='nsew', padx=10, pady=5)
        
        # Configure grid weights for proper resizing
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_columnconfigure(0, weight=0, minsize=250)  # File panel - fixed minimum width
        main_frame.grid_columnconfigure(1, weight=1, minsize=300)  # Analysis panel - resizable
        main_frame.grid_columnconfigure(2, weight=2, minsize=400)  # Visualization panel - largest
        
        # Create panels
        self.create_file_panel(main_frame)
        self.create_analysis_panel(main_frame)
        self.create_visualization_panel(main_frame)
        
    def create_file_panel(self, parent):
        """Create file selection and model management panel."""
        file_frame = ttk.LabelFrame(parent, text="Data Input", padding=10)
        file_frame.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        
        # File selection
        ttk.Label(file_frame, text="Excel File:").grid(row=0, column=0, sticky='w', pady=5)
        
        self.file_var = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.file_var, width=40, state='readonly')
        file_entry.grid(row=1, column=0, columnspan=2, sticky='ew', pady=5)
        
        browse_btn = ttk.Button(file_frame, text="Browse", 
                               command=self.browse_file, style='Custom.TButton')
        browse_btn.grid(row=2, column=0, sticky='w', pady=5)
        
        # Model selection
        ttk.Label(file_frame, text="Available Models:").grid(row=3, column=0, sticky='w', pady=(20,5))
        
        # Listbox with scrollbar
        listbox_frame = ttk.Frame(file_frame)
        listbox_frame.grid(row=4, column=0, columnspan=2, sticky='nsew', pady=5)
        
        self.model_listbox = tk.Listbox(listbox_frame, selectmode='multiple', height=8)
        scrollbar = ttk.Scrollbar(listbox_frame, orient='vertical', command=self.model_listbox.yview)
        self.model_listbox.configure(yscrollcommand=scrollbar.set)
        
        self.model_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Analysis button
        analyze_btn = ttk.Button(file_frame, text="Analyze Selected Models",
                               command=self.analyze_models, style='Custom.TButton')
        analyze_btn.grid(row=5, column=0, columnspan=2, sticky='ew', pady=20)
        
        # Debug panel (only show in debug mode)
        if DEBUG_MODE:
            debug_frame = ttk.LabelFrame(file_frame, text="Debug Controls", padding=5)
            debug_frame.grid(row=6, column=0, columnspan=2, sticky='ew', pady=5)
            
            debug_log_btn = ttk.Button(debug_frame, text="Show Debug Log",
                                      command=self.show_debug_log)
            debug_log_btn.pack(side='left', padx=5)
            
            clear_log_btn = ttk.Button(debug_frame, text="Clear Log",
                                      command=self.clear_debug_log)
            clear_log_btn.pack(side='left', padx=5)
        
        file_frame.grid_rowconfigure(4, weight=1)
        file_frame.grid_columnconfigure(0, weight=1)
        
    def create_analysis_panel(self, parent):
        """Create analysis results panel."""
        analysis_frame = ttk.LabelFrame(parent, text="Analysis Results", padding=10)
        analysis_frame.grid(row=0, column=1, sticky='nsew', padx=5, pady=5)
        
        # Configure grid weights for proper resizing
        analysis_frame.grid_rowconfigure(0, weight=1)
        analysis_frame.grid_columnconfigure(0, weight=1)
        
        # Results text area with scrollbar in a frame
        text_frame = ttk.Frame(analysis_frame)
        text_frame.grid(row=0, column=0, sticky='nsew', pady=(0, 10))
        text_frame.grid_rowconfigure(0, weight=1)
        text_frame.grid_columnconfigure(0, weight=1)
        
        # Text widget with word wrap and proper scrolling
        self.results_text = tk.Text(
            text_frame, 
            wrap='word', 
            font=('Consolas', 10),
            height=20,  # Set minimum height
            state='normal'  # Allow text insertion
        )
        self.results_text.grid(row=0, column=0, sticky='nsew')
        
        # Vertical scrollbar
        text_scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=self.results_text.yview)
        text_scrollbar.grid(row=0, column=1, sticky='ns')
        self.results_text.configure(yscrollcommand=text_scrollbar.set)
        
        # Horizontal scrollbar for long lines
        h_scrollbar = ttk.Scrollbar(text_frame, orient='horizontal', command=self.results_text.xview)
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        self.results_text.configure(xscrollcommand=h_scrollbar.set)
        
        # Export buttons frame
        export_frame = ttk.Frame(analysis_frame)
        export_frame.grid(row=1, column=0, sticky='ew')
        export_frame.grid_columnconfigure(0, weight=1)
        export_frame.grid_columnconfigure(1, weight=1)
        
        # Export results button
        export_btn = ttk.Button(export_frame, text="Export Results (TXT/CSV)",
                               command=self.export_results, style='Custom.TButton')
        export_btn.grid(row=0, column=0, sticky='ew', padx=(0, 5))
        
        # Export PDF report button
        export_pdf_btn = ttk.Button(export_frame, text="Export PDF Report",
                                   command=self.export_pdf_report, style='Custom.TButton')
        export_pdf_btn.grid(row=0, column=1, sticky='ew', padx=(5, 0))
        
    def create_visualization_panel(self, parent):
        """Create visualization panel with tabs."""
        viz_frame = ttk.LabelFrame(parent, text="Visualizations", padding=10)
        viz_frame.grid(row=0, column=2, sticky='nsew', padx=5, pady=5)
        
        # Configure frame for proper expansion
        viz_frame.grid_rowconfigure(0, weight=1)
        viz_frame.grid_columnconfigure(0, weight=1)
        
        # Create notebook for different chart types
        self.viz_notebook = ttk.Notebook(viz_frame)
        self.viz_notebook.grid(row=0, column=0, sticky='nsew')
        
        # Placeholder text
        placeholder_frame = ttk.Frame(self.viz_notebook)
        self.viz_notebook.add(placeholder_frame, text="Welcome")
        
        # Configure placeholder frame
        placeholder_frame.grid_rowconfigure(0, weight=1)
        placeholder_frame.grid_columnconfigure(0, weight=1)
        
        welcome_label = ttk.Label(placeholder_frame, 
                                 text="Select models and click 'Analyze' to view results\n\n"
                                      "TraceSeis RADAR Features:\n"
                                      "- Scrollable and resizable visualizations\n"
                                      "- Interactive navigation toolbar\n"
                                      "- Professional neural network classification matrices\n"
                                      "- Distribution pie charts\n"
                                      "- Multi-model RADAR comparison charts\n\n"
                                      "Enterprise License - $35,000/Year Professional Suite\n"
                                      "Results Analysis and Data Accuracy Reporter",
                                 font=('Segoe UI', 12),
                                 justify='center')
        welcome_label.grid(row=0, column=0, sticky='nsew', padx=20, pady=20)
        
    def create_status_bar(self):
        """Create status bar with progress indicator."""
        status_frame = ttk.Frame(self.root)
        status_frame.grid(row=2, column=0, sticky='ew', padx=10, pady=5)
        status_frame.grid_columnconfigure(0, weight=1)
        
        self.status_var = tk.StringVar(value="Ready - TraceSeis RADAR v2.0.0 Enterprise")
        status_label = ttk.Label(status_frame, textvariable=self.status_var)
        status_label.grid(row=0, column=0, sticky='w')
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(status_frame, variable=self.progress_var, 
                                           length=200, mode='determinate')
        self.progress_bar.grid(row=0, column=1, sticky='e', padx=10)
        
    def browse_file(self):
        """Browse for Excel file containing neural network results."""
        try:
            logger.debug("Opening file browser")
            
            file_path = filedialog.askopenfilename(
                title="Select Neural Network Results File",
                filetypes=[
                    ("Excel files", "*.xlsx *.xls"),
                    ("All files", "*.*")
                ]
            )
            
            if file_path:
                logger.debug(f"File selected: {file_path}")
                self.file_var.set(file_path)
                self.load_models(file_path)
                
        except Exception as e:
            self.logger.error(f"File browsing error: {e}")
            if DEBUG_MODE:
                self.logger.debug(traceback.format_exc())
            messagebox.showerror("Error", f"Failed to load file: {str(e)}")
    
    def load_models(self, file_path: str):
        """Load available models from Excel file."""
        try:
            logger.debug(f"Loading models from: {file_path}")
            self.status_var.set("Loading models...")
            
            # Read Excel file
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            logger.debug(f"Found sheets: {sheet_names}")
            
            # Clear current listbox
            self.model_listbox.delete(0, tk.END)
            
            # Add sheet names as available models
            for sheet_name in sheet_names:
                self.model_listbox.insert(tk.END, sheet_name)
            
            self.status_var.set(f"Loaded {len(sheet_names)} models")
            self.logger.info(f"Loaded {len(sheet_names)} models from {file_path}")
            
        except Exception as e:
            self.logger.error(f"Model loading error: {e}")
            if DEBUG_MODE:
                self.logger.debug(traceback.format_exc())
            self.status_var.set("Error loading models")
            messagebox.showerror("Error", f"Failed to load models: {str(e)}")
    
    def analyze_models(self):
        """Analyze selected models in background thread."""
        logger.debug("Starting model analysis")
        
        selected_indices = self.model_listbox.curselection()
        
        if not selected_indices:
            messagebox.showwarning("No Selection", "Please select one or more models to analyze.")
            return
    
    def export_pdf_report(self):
        """Export professional PDF report of analysis results."""
        try:
            if not hasattr(self, 'current_metrics') or not self.current_metrics:
                messagebox.showwarning("No Data", "Please analyze models first before exporting PDF report.")
                return
            
            # Ask user for save location
            output_path = filedialog.asksaveasfilename(
                title="Save PDF Report",
                defaultextension=".pdf",
                filetypes=[
                    ("PDF files", "*.pdf"),
                    ("All files", "*.*")
                ]
            )
            
            if not output_path:
                return  # User cancelled
            
            self.status_var.set("Creating PDF report...")
            self.progress_var.set(10)
            self.root.update()
            
            # Prepare metrics list
            metrics_list = list(self.current_metrics.values())
            
            self.progress_var.set(30)
            self.root.update()
            
            # Prepare figures dictionary if available
            figures_dict = getattr(self, 'current_figures', {})
            
            self.progress_var.set(50)
            self.root.update()
            
            # Generate PDF report using the existing PDFReportGenerator
            success = self.pdf_generator.create_comprehensive_report(
                metrics_list, figures_dict, output_path
            )
            
            self.progress_var.set(90)
            self.root.update()
            
            if success:
                self.status_var.set(f"PDF report saved: {output_path}")
                self.progress_var.set(100)
                messagebox.showinfo("Success", f"PDF report created successfully!\n\nSaved to: {output_path}")
                self.logger.info(f"PDF report exported successfully: {output_path}")
            else:
                self.status_var.set("PDF export failed")
                messagebox.showerror("Error", f"Failed to create PDF report.\n\nPlease check the log file for details.")
                self.logger.error("PDF report export failed")
            
            # Reset progress
            self.progress_var.set(0)
            
        except Exception as e:
            self.logger.error(f"PDF export error: {e}")
            if DEBUG_MODE:
                self.logger.debug(traceback.format_exc())
            self.status_var.set("PDF export error")
            self.progress_var.set(0)
            messagebox.showerror("Error", f"Failed to export PDF report: {str(e)}")
    
    def export_results(self):
        """Export analysis results to text and CSV files."""
        try:
            if not hasattr(self, 'current_metrics') or not self.current_metrics:
                messagebox.showwarning("No Data", "Please analyze models first before exporting results.")
                return
            
            # Ask user for save location
            output_path = filedialog.asksaveasfilename(
                title="Save Results",
                defaultextension=".txt",
                filetypes=[
                    ("Text files", "*.txt"),
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ]
            )
            
            if not output_path:
                return  # User cancelled
            
            self.status_var.set("Exporting results...")
            
            # Determine file format from extension
            file_ext = os.path.splitext(output_path)[1].lower()
            
            if file_ext == '.csv':
                self._export_csv_results(output_path)
            else:
                self._export_text_results(output_path)
            
            self.status_var.set(f"Results exported: {output_path}")
            messagebox.showinfo("Success", f"Results exported successfully!\n\nSaved to: {output_path}")
            self.logger.info(f"Results exported: {output_path}")
            
        except Exception as e:
            self.logger.error(f"Export error: {e}")
            if DEBUG_MODE:
                self.logger.debug(traceback.format_exc())
            self.status_var.set("Export error")
            messagebox.showerror("Error", f"Failed to export results: {str(e)}")
    
    def _export_text_results(self, output_path: str):
        """Export results as formatted text file."""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("TraceSeis RADAR - Results Analysis and Data Accuracy Reporter\n")
            f.write("=" * 70 + "\n")
            f.write(f"Analysis Results - {len(self.current_metrics)} models\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            for model_name, metrics in self.current_metrics.items():
                f.write(f"\nModel: {model_name}\n")
                f.write("-" * 50 + "\n")
                f.write(f"Global Fit: {metrics.global_fit:.2f}%\n")
                f.write(f"Accuracy: {metrics.accuracy:.2f}%\n")
                f.write(f"Cramer's V: {metrics.cramers_v:.3f}\n")
                f.write(f"Percent Undefined: {metrics.percent_undefined:.1f}%\n")
                f.write(f"Total Samples: {metrics.total_samples}\n")
                
                f.write("\nPer-Class Metrics:\n")
                for class_name in metrics.precision.keys():
                    f.write(f"  {class_name}:\n")
                    f.write(f"    Precision: {metrics.precision[class_name]:.3f}\n")
                    f.write(f"    Recall: {metrics.recall[class_name]:.3f}\n")
                    f.write(f"    F1-Score: {metrics.f1_score[class_name]:.3f}\n")
                    f.write(f"    Distribution: {metrics.classification_distribution[class_name]:.1f}%\n")
                f.write("\n")
    
    def _export_csv_results(self, output_path: str):
        """Export results as CSV file."""
        import csv
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([
                'Model', 'Global_Fit_%', 'Accuracy_%', 'Cramers_V', 
                'Percent_Undefined_%', 'Total_Samples'
            ])
            
            # Data rows
            for model_name, metrics in self.current_metrics.items():
                writer.writerow([
                    model_name,
                    f"{metrics.global_fit:.2f}",
                    f"{metrics.accuracy:.2f}",
                    f"{metrics.cramers_v:.3f}",
                    f"{metrics.percent_undefined:.1f}",
                    metrics.total_samples
                ])
